import collections
import datetime

import openpyxl
from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.utils import get_column_letter

from .models import Robot


def get_orders(request):
    start_datetime = timezone.now() - datetime.timedelta(days=7)
    data = Robot.objects.filter(created__gte=start_datetime).values('model', 'version').annotate(count=Count('id'))

    grouped_data = collections.defaultdict(list)
    for item in data:
        grouped_data[item['model']].append((item['version'], item['count']))

    workbook = openpyxl.Workbook()
    del workbook['Sheet']
    header = ['Модель', 'Версия', 'Количество за неделю']

    for model, versions_data in grouped_data.items():
        worksheet = workbook.create_sheet(title=model)

        for col_num, column_title in enumerate(header, 1):
            column_letter = get_column_letter(col_num)
            worksheet[f'{column_letter}1'] = column_title

        for row_num, (version, count) in enumerate(versions_data, 2):
            worksheet.cell(row=row_num, column=1, value=model)
            worksheet.cell(row=row_num, column=2, value=version)
            worksheet.cell(row=row_num, column=3, value=count)

        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

    if data:
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=last_7_days_robots.xlsx'
        workbook.save(response)
    else:
        response = HttpResponse('<h2>За последние 7 дней роботов не создавалось</h2>')

    return response
